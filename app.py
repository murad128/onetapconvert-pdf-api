import os, io, base64, json, threading, time, urllib.request
from flask import Flask, request, jsonify
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

# ── PDF type detection ────────────────────────────────────────────────────────
def detect_pdf_type(pdf_bytes):
    import pdfplumber
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        total_chars = sum(len((p.extract_text() or '').strip()) for p in pdf.pages[:3])
    return 'text' if total_chars > 80 else 'scanned'

# ── pdfplumber extraction (primary) ──────────────────────────────────────────
def extract_with_pdfplumber(pdf_bytes):
    import pdfplumber
    all_tables = []  # list of (table_rows)

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in (tables or []):
                rows = []
                for row in table:
                    clean = [str(c or '').replace('\n', ' ').strip() for c in row]
                    non_empty = [c for c in clean if c]
                    if not non_empty:
                        continue
                    # Skip rows where all content is in a single cell (header/address blocks)
                    if len(non_empty) == 1 and len(clean) > 2:
                        continue
                    rows.append(clean)
                if rows:
                    all_tables.append(rows)

    if not all_tables:
        # Fallback: extract all text as single-column rows
        fallback_rows = []
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                for line in text.splitlines():
                    line = line.strip()
                    if line:
                        fallback_rows.append([line])
        return [fallback_rows] if fallback_rows else []

    # Normalize each table's rows to its own max col count
    result = []
    for table in all_tables:
        max_cols = max(len(r) for r in table)
        normalized = [(r + [''] * max_cols)[:max_cols] for r in table]
        result.append(normalized)

    return result

# ── XLSX export ───────────────────────────────────────────────────────────────
def write_sheet(ws, rows):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style='thin', color='CCCCCC')
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    no_fill = PatternFill()

    max_cols = max(len(r) for r in rows) if rows else 0
    col_max = {}

    # Detect key-value table (2 cols, no obvious header)
    is_kv_table = max_cols == 2

    # Detect if first row is a real header (all caps or common header words)
    def is_header_row(row):
        vals = [str(v).strip() for v in row if str(v).strip()]
        if not vals: return False
        header_words = {'no', 'oem', 'fmsi', 'description', 'qty', 'price', 'amount', 'item', '#'}
        return any(v.lower() in header_words or v.isupper() for v in vals)

    has_header = not is_kv_table and is_header_row(rows[0]) if rows else False

    for r_idx, row in enumerate(rows):
        is_hdr = has_header and r_idx == 0
        for c_idx in range(max_cols):
            val = row[c_idx] if c_idx < len(row) else ''
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            cell.border = bdr
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            col_max[c_idx + 1] = max(col_max.get(c_idx + 1, 8), len(str(val)))

            if is_hdr:
                cell.font = Font(bold=True, color='FFFFFF', size=10)
                cell.fill = hdr_fill
            elif is_kv_table:
                # Key-value: left col normal, right col bold
                cell.font = Font(bold=(c_idx == 1), size=10)
                cell.fill = alt_fill if r_idx % 2 == 0 else no_fill
            else:
                cell.font = Font(size=10)
                cell.fill = alt_fill if r_idx % 2 == 0 else no_fill

    for ci, ml in col_max.items():
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 3, 50)

    if has_header and rows:
        ws.freeze_panes = 'A2'

def tables_to_xlsx(all_tables):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for i, rows in enumerate(all_tables, 1):
        ws = wb.create_sheet(title=f'Table {i}')
        write_sheet(ws, rows)

    if not wb.sheetnames:
        return None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def rows_to_xlsx(all_rows):
    """Single-sheet export (kept for compatibility)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    write_sheet(ws, all_rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── Main route ────────────────────────────────────────────────────────────────
@app.route('/convert', methods=['POST'])
def convert():
    try:
        data = request.get_json()
        file_b64 = data.get('fileBase64', '')
        file_name = data.get('fileName', 'input.pdf')
        if not file_b64:
            return jsonify({'error': 'No file provided'}), 400

        pdf_bytes = base64.b64decode(file_b64)
        pdf_type = detect_pdf_type(pdf_bytes)
        out_name = file_name.rsplit('.', 1)[0] + '.xlsx'
        warning = None

        if pdf_type == 'text':
            all_tables = extract_with_pdfplumber(pdf_bytes)
            if not all_tables:
                return jsonify({'error': 'No content could be extracted from this PDF.'}), 422
            xlsx_bytes = tables_to_xlsx(all_tables)
            method = 'pdfplumber'
        else:
            all_tables = extract_with_pdfplumber(pdf_bytes)
            warning = 'Scanned PDF detected. Results may be lower accuracy.'
            if not all_tables:
                return jsonify({'error': 'No content could be extracted. This appears to be a scanned image PDF.'}), 422
            xlsx_bytes = tables_to_xlsx(all_tables)
            method = 'pdfplumber-scanned'

        if not xlsx_bytes:
            return jsonify({'error': 'Failed to generate Excel file.'}), 500

        xlsx_b64 = base64.b64encode(xlsx_bytes).decode('utf-8')
        return jsonify({
            'base64': xlsx_b64,
            'fileName': out_name,
            'pdfType': pdf_type,
            'method': method,
            'warning': warning
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/office-to-pdf', methods=['POST'])
def office_to_pdf():
    """Convert Word/Excel/PPT to PDF using LibreOffice."""
    import subprocess, tempfile, os
    try:
        data = request.get_json()
        file_b64 = data.get('fileBase64', '')
        file_name = data.get('fileName', 'input.docx')
        if not file_b64:
            return jsonify({'error': 'No file provided'}), 400

        file_bytes = base64.b64decode(file_b64)

        with tempfile.TemporaryDirectory() as tmpdir:
            in_path = os.path.join(tmpdir, file_name)
            with open(in_path, 'wb') as f:
                f.write(file_bytes)

            # Run LibreOffice headless conversion
            # Use pdf:writer_pdf_Export with scale settings for PPTX
            ext = os.path.splitext(file_name)[1].lower()
            if ext in ['.pptx', '.ppt']:
                convert_to = 'pdf:impress_pdf_Export:{"ReduceImageResolution":{"type":"boolean","value":"false"},"IsSkipEmptyPages":{"type":"boolean","value":"false"}}'
            else:
                convert_to = 'pdf'

            result = subprocess.run([
                'libreoffice', '--headless',
                '--convert-to', convert_to,
                '--outdir', tmpdir, in_path
            ], capture_output=True, text=True, timeout=60)

            if result.returncode != 0:
                return jsonify({'error': 'Conversion failed: ' + result.stderr[:200]}), 500

            # Find output PDF
            out_name = os.path.splitext(file_name)[0] + '.pdf'
            out_path = os.path.join(tmpdir, out_name)
            if not os.path.exists(out_path):
                # Try to find any PDF in tmpdir
                pdfs = [f for f in os.listdir(tmpdir) if f.endswith('.pdf')]
                if not pdfs:
                    return jsonify({'error': 'Output PDF not found'}), 500
                out_path = os.path.join(tmpdir, pdfs[0])
                out_name = pdfs[0]

            with open(out_path, 'rb') as f:
                pdf_bytes = f.read()

            return jsonify({
                'base64': base64.b64encode(pdf_bytes).decode('utf-8'),
                'fileName': out_name
            })

    except subprocess.TimeoutExpired:
        return jsonify({'error': 'Conversion timed out'}), 504
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/feedback', methods=['POST'])
def feedback():
    import urllib.request as ur
    try:
        data = request.get_json()
        message = data.get('message', '').strip()
        if not message:
            return jsonify({'error': 'No message'}), 400

        BOT_TOKEN = '8798510190:AAEwnO5ZjICKqL6MTiLlcCaqBQJ1aTJUO4A'
        CHAT_ID = '1871988010'

        text = f"🔔 *OneTapConvert Feedback*\n\n📝 *Mesaj:* {message}"
        if data.get('email'):
            text += f"\n📧 *Email:* {data['email']}"
        if data.get('url'):
            text += f"\n🔗 *Səhifə:* {data['url']}"

        req = ur.Request(
            f'https://api.telegram.org/bot{BOT_TOKEN}/sendMessage',
            data=json.dumps({'chat_id': CHAT_ID, 'text': text, 'parse_mode': 'Markdown'}).encode(),
            headers={'Content-Type': 'application/json'}
        )
        ur.urlopen(req, timeout=10)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

# ── United Center CMS ─────────────────────────────────────────────────────────
import hashlib

UC_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uc_content.json')
UC_ADMIN_HASH = hashlib.sha256('Admin12345'.encode()).hexdigest()

def uc_read():
    try:
        with open(UC_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return None

def uc_write(data):
    with open(UC_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

DEFAULT_CONTENT = {
  "contact": {"phone": "+994 50 204 64 30", "phoneRaw": "+994502046430", "address": "Səttar Bəhlulzadə 157", "whatsapp": "994502046430", "email": "", "hoursWeekday": "09:00 – 18:00", "hoursSaturday": "09:00 – 14:00"},
  "hero": {"title": "Uşağınızın potensialını", "titleSpan": "birlikdə açaq", "subtitle": "United Center-də hər uşaq xüsusidir. Autizm dəstəyi, nitq terapiyası, psixoloji yardım və fərdi inkişaf proqramları ilə yanınızdayıq."},
  "stats": {"children": "100", "childrenSuffix": "+", "experience": "5", "experienceSuffix": "+", "specialists": "8", "specialistsSuffix": "+", "satisfaction": "95", "satisfactionSuffix": "%"},
  "team": [
    {"name": "Dr. Aynur Məmmədova", "role": "Baş Psixoloq", "desc": "10 ildən artıq uşaq psixologiyası sahəsində təcrübəyə malikdir.", "initials": "AM", "color1": "#17B8A6", "color2": "#0e9e8e"},
    {"name": "Gülnar Əliyeva", "role": "Nitq Terapevti / Loqoped", "desc": "Nitq pozğunluqları və kommunikasiya inkişafı üzrə mütəxəssis.", "initials": "GƏ", "color1": "#4CAF50", "color2": "#388E3C"},
    {"name": "Nigar Hüseynova", "role": "Xüsusi Müəllim", "desc": "Fərdi öyrənmə planları ilə uşaqların potensialını açır.", "initials": "NH", "color1": "#2196F3", "color2": "#1565C0"},
    {"name": "Rauf Quliyev", "role": "Davranış Mütəxəssisi", "desc": "ABA terapiyası üzrə sertifikatlaşmış mütəxəssis.", "initials": "RQ", "color1": "#FF9800", "color2": "#E65100"},
    {"name": "Sevinc İsmayılova", "role": "Uşaq İnkişaf Mütəxəssisi", "desc": "Erkən müdaxilə və kompleks inkişaf proqramları üzrə təcrübəli.", "initials": "Sİ", "color1": "#9C27B0", "color2": "#6A1B9A"}
  ],
  "services": [
    {"icon": "🧩", "name": "Autizm Dəstəyi", "desc": "Fərdi rehabilitasiya proqramları"},
    {"icon": "🗣️", "name": "Nitq Terapiyası", "desc": "Loqoped ilə nitq inkişafı"},
    {"icon": "🧠", "name": "Psixoloji Dəstək", "desc": "Uşaq psixoloqu ilə iş"},
    {"icon": "📚", "name": "Xüsusi Təhsil", "desc": "Fərdi öyrənmə planları"},
    {"icon": "🌱", "name": "Uşaq İnkişafı", "desc": "Kompleks inkişaf proqramları"},
    {"icon": "🎯", "name": "Davranış Terapiyası", "desc": "ABA əsaslı yanaşma"},
    {"icon": "🤲", "name": "Sensor İnteqrasiya", "desc": "Hiss sisteminin inkişafı"},
    {"icon": "👨‍👩‍👧", "name": "Valideyn Konsultasiyası", "desc": "Evdə dəstək strategiyaları"},
    {"icon": "📋", "name": "Fərdi Proqramlar", "desc": "Hər uşaq üçün xüsusi plan"}
  ],
  "testimonials": [
    {"name": "Leyla X.", "info": "Valideyn", "initials": "LX", "color": "#17B8A6", "text": "Oğlumun nitqi 3 ayda əhəmiyyətli dərəcədə inkişaf etdi.", "stars": 5},
    {"name": "Kamran M.", "info": "Valideyn", "initials": "KM", "color": "#4CAF50", "text": "Qızımız burada çox irəlilədi. Mütəxəssislər həm peşəkar, həm də mehriban.", "stars": 5},
    {"name": "Aynur R.", "info": "Valideyn", "initials": "AR", "color": "#2196F3", "text": "United Center-in fərdi yanaşması bizim üçün çox böyük fərq yaratdı.", "stars": 5}
  ]
}

@app.route('/uc/content', methods=['GET'])
def uc_get_content():
    data = uc_read()
    if data:
        return jsonify(data)
    return jsonify(DEFAULT_CONTENT)

@app.route('/uc/content', methods=['PUT'])
def uc_put_content():
    auth = request.headers.get('X-Admin-Password', '')
    if hashlib.sha256(auth.encode()).hexdigest() != UC_ADMIN_HASH:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data'}), 400
        uc_write(data)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── New PDF Tools ─────────────────────────────────────────────────────────────

@app.route('/repair-pdf', methods=['POST'])
def repair_pdf():
    try:
        import pikepdf
        data = request.get_json()
        file_b64 = data.get('fileBase64', '')
        file_name = data.get('fileName', 'input.pdf')
        if not file_b64:
            return jsonify({'error': 'No file provided'}), 400
        pdf_bytes = base64.b64decode(file_b64)
        buf = io.BytesIO(pdf_bytes)
        out_buf = io.BytesIO()
        with pikepdf.open(buf, suppress_warnings=True) as pdf:
            pdf.save(out_buf)
        out_buf.seek(0)
        out_bytes = out_buf.read()
        out_name = file_name.rsplit('.', 1)[0] + '-repaired.pdf'
        return jsonify({'base64': base64.b64encode(out_bytes).decode(), 'fileName': out_name})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/pdf-to-pdfa', methods=['POST'])
def pdf_to_pdfa():
    import subprocess, tempfile
    try:
        data = request.get_json()
        file_b64 = data.get('fileBase64', '')
        file_name = data.get('fileName', 'input.pdf')
        if not file_b64:
            return jsonify({'error': 'No file provided'}), 400
        pdf_bytes = base64.b64decode(file_b64)
        with tempfile.TemporaryDirectory() as tmp:
            in_path = os.path.join(tmp, file_name)
            out_name = file_name.rsplit('.', 1)[0] + '-pdfa.pdf'
            out_path = os.path.join(tmp, out_name)
            with open(in_path, 'wb') as f:
                f.write(pdf_bytes)
            result = subprocess.run([
                'gs', '-dPDFA=2', '-dBATCH', '-dNOPAUSE',
                '-sColorConversionStrategy=RGB',
                '-sDEVICE=pdfwrite',
                '-dPDFACompatibilityPolicy=2',
                f'-sOutputFile={out_path}', in_path
            ], capture_output=True, text=True, timeout=60)
            if result.returncode != 0 or not os.path.exists(out_path):
                # Fallback: just re-save with pikepdf
                import pikepdf
                with pikepdf.open(in_path) as pdf:
                    pdf.save(out_path)
            with open(out_path, 'rb') as f:
                out_bytes = f.read()
        return jsonify({'base64': base64.b64encode(out_bytes).decode(), 'fileName': out_name})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/compare-pdf', methods=['POST'])
def compare_pdf():
    try:
        import pdfplumber, difflib
        data = request.get_json()
        b64a = data.get('file1Base64', '')
        b64b = data.get('file2Base64', '')
        if not b64a or not b64b:
            return jsonify({'error': 'Two PDF files required'}), 400
        def extract_pages(b64):
            pdf_bytes = base64.b64decode(b64)
            pages = []
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page in pdf.pages:
                    pages.append((page.extract_text() or '').strip())
            return pages
        pages_a = extract_pages(b64a)
        pages_b = extract_pages(b64b)
        differences = []
        max_pages = max(len(pages_a), len(pages_b))
        for i in range(max_pages):
            ta = pages_a[i] if i < len(pages_a) else ''
            tb = pages_b[i] if i < len(pages_b) else ''
            if ta != tb:
                diff = list(difflib.unified_diff(ta.splitlines(), tb.splitlines(), lineterm=''))
                removed = '\n'.join(l[1:] for l in diff if l.startswith('-') and not l.startswith('---'))
                added = '\n'.join(l[1:] for l in diff if l.startswith('+') and not l.startswith('+++'))
                differences.append({'page': i + 1, 'removed': removed[:500], 'added': added[:500]})
        return jsonify({'differences': differences, 'totalPages': max_pages})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/html-to-pdf', methods=['POST'])
def html_to_pdf():
    import subprocess, tempfile, re, urllib.request as ur
    try:
        data = request.get_json()
        html_content = data.get('html', '')
        url = data.get('url', '')
        if not html_content and not url:
            return jsonify({'error': 'html or url required'}), 400

        # Skip weasyprint (version incompatibility), use fpdf2 directly
        pass

        # Fallback: fetch URL content if needed
        if url and not html_content:
            try:
                req = ur.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                html_content = ur.urlopen(req, timeout=20).read().decode('utf-8', errors='ignore')
            except Exception as e:
                return jsonify({'error': f'Could not fetch URL: {e}'}), 400
            fname = 'webpage.pdf'
        else:
            fname = 'converted.pdf'

        # Fallback: build PDF from extracted text using fpdf2 or pikepdf
        try:
            from fpdf import FPDF
            text = re.sub(r'<[^>]+>', ' ', html_content)
            text = re.sub(r'\s+', ' ', text).strip()
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Helvetica', size=11)
            pdf.set_left_margin(15)
            pdf.set_right_margin(15)
            for line in text.split('. ')[:200]:
                line = line.strip()
                if line:
                    try:
                        pdf.multi_cell(0, 8, line.encode('latin-1', errors='replace').decode('latin-1'))
                    except:
                        pass
            out_bytes = pdf.output()
            if isinstance(out_bytes, str):
                out_bytes = out_bytes.encode('latin-1')
            return jsonify({'base64': base64.b64encode(out_bytes).decode(), 'fileName': fname})
        except ImportError:
            pass

        # Last resort: minimal PDF with text
        try:
            import pikepdf
            from pikepdf import Pdf, Page, Dictionary, Name, Array
            text = re.sub(r'<[^>]+>', ' ', html_content)[:2000]
            # Create simple PDF with text content note
            pdf = Pdf.new()
            page = pikepdf.Page(pikepdf.Dictionary(
                Type=pikepdf.Name.Page,
                MediaBox=pikepdf.Array([0, 0, 612, 792])
            ))
            pdf.pages.append(page)
            buf = io.BytesIO()
            pdf.save(buf)
            buf.seek(0)
            return jsonify({'base64': base64.b64encode(buf.read()).decode(), 'fileName': fname})
        except:
            return jsonify({'error': 'HTML to PDF conversion requires weasyprint or fpdf2. Install fpdf2 on the server.'}), 500

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/ocr-pdf', methods=['POST'])
def ocr_pdf():
    try:
        data = request.get_json()
        file_b64 = data.get('fileBase64', '')
        file_name = data.get('fileName', 'input.pdf')
        lang = data.get('lang', 'eng')
        if not file_b64:
            return jsonify({'error': 'No file provided'}), 400
        pdf_bytes = base64.b64decode(file_b64)
        # Try pytesseract + pdf2image
        try:
            from pdf2image import convert_from_bytes
            import pytesseract
            images = convert_from_bytes(pdf_bytes, dpi=200)
            texts = []
            for img in images:
                t = pytesseract.image_to_string(img, lang=lang)
                texts.append(t)
            full_text = '\n\n--- Page Break ---\n\n'.join(texts)
            return jsonify({'text': full_text, 'pages': len(images)})
        except ImportError:
            # Fallback: use pdfplumber text extraction
            import pdfplumber
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                texts = [(page.extract_text() or '') for page in pdf.pages]
            full_text = '\n\n--- Page Break ---\n\n'.join(texts)
            if not full_text.strip():
                return jsonify({'error': 'No text found. This appears to be a scanned image PDF. OCR requires pytesseract to be installed on the server.'}), 422
            return jsonify({'text': full_text, 'pages': len(texts)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/ocr-pdf-upload', methods=['POST'])
def ocr_pdf_upload():
    """OCR via multipart/form-data — avoids base64 overhead on mobile"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        f = request.files['file']
        lang = request.form.get('lang', 'eng')
        pdf_bytes = f.read()
        file_name = f.filename or 'input.pdf'

        try:
            from pdf2image import convert_from_bytes
            import pytesseract
            images = convert_from_bytes(pdf_bytes, dpi=200)
            texts = []
            for img in images:
                t = pytesseract.image_to_string(img, lang=lang)
                texts.append(t)
            full_text = '\n\n--- Page Break ---\n\n'.join(texts)
            return jsonify({'text': full_text, 'pages': len(images)})
        except ImportError:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                texts = [(page.extract_text() or '') for page in pdf.pages]
            full_text = '\n\n--- Page Break ---\n\n'.join(texts)
            if not full_text.strip():
                return jsonify({'error': 'No text found in PDF. OCR requires pytesseract on the server.'}), 422
            return jsonify({'text': full_text, 'pages': len(texts)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ── Keep-alive ────────────────────────────────────────────────────────────────
def keep_alive():
    time.sleep(60)
    while True:
        try:
            host = os.environ.get('RAILWAY_PUBLIC_DOMAIN') or 'onetapconvert-pdf-api.onrender.com'
            urllib.request.urlopen(f'https://{host}/health', timeout=10)
        except:
            pass
        time.sleep(600)

threading.Thread(target=keep_alive, daemon=True).start()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
